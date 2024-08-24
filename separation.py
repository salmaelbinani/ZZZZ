import sys
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import numpy as np
import re

def generate_ds_general(sn_value, df_mmsta):
    if pd.isna(sn_value):
        return ''
    
    start_rows = df_mmsta[df_mmsta['Material'] == sn_value]
    if start_rows.empty:
        return ''
    
    start_index = start_rows.index[0]
    ds_general_parts = []
    
    # Concaténer les valeurs verticalement à partir de la ligne de départ
    for i in range(start_index + 1, len(df_mmsta)):
        cell_value = df_mmsta.iloc[i]['Material']
        if pd.isna(cell_value):
            continue
        cell_value_str = str(cell_value)
        if cell_value_str.startswith('S00'):
            break
        if df_mmsta.iloc[i]['Material Description'].startswith('LT Single Wire'):
            component_quantity = df_mmsta.iloc[i]['Component quantity']
            ds_general_parts.append(f"{cell_value_str} ({component_quantity})")
        else:
            ds_general_parts.append(cell_value_str)
    
    return ' / '.join(ds_general_parts)

def separate_files(mmsta_file):
    try:
        df_mmsta = pd.read_excel(mmsta_file, sheet_name='MMSTA')
    except Exception as e:
        print(f"Erreur lors de la lecture du fichier MMSTA: {e}")
        sys.exit(1)

    def update_pn(row):
        level = extract_level(row['Level'])
        if level == 0:
            return row['Material Description']
        return None

    def extract_level(level_str):
        match = re.search(r'\d+', level_str)
        return int(match.group()) if match else None

    df_mmsta['PN'] = df_mmsta.apply(update_pn, axis=1)
    df_mmsta['PN'] = df_mmsta['PN'].fillna(method='ffill')

    def count_asterisks(value):
        return value.count('*')

    df_mmsta['asterisk_count'] = df_mmsta['Level'].apply(count_asterisks)
    max_asterisks = df_mmsta['asterisk_count'].max()
    n = max_asterisks + 1

    df_1 = df_mmsta[(df_mmsta['Level'] != 0) & (~df_mmsta['Level'].astype(str).str.endswith(str(max_asterisks))) & (df_mmsta['Material Type'] == 'YSFG')].copy()
    for i in range(1, n):
        df_1[f'SN{i}'] = np.nan
        df_1[f'DSN{i}'] = np.nan

    df_1.drop(columns=['asterisk_count'], inplace=True)

    def update_sn(row, level_indicator):
        if row['Level'] == level_indicator:
            return row['Material']
        return np.nan

    def update_and_fill_sn(df_1, sn_column, level_indicator):
        df_1.loc[:, sn_column] = df_1.apply(lambda row: update_sn(row, level_indicator), axis=1)
        df_1[sn_column] = df_1[sn_column].fillna(method='ffill')

    def generate_level(n):
        return '*' * (n + 1) + str(n + 1)

    for i in range(n-1):
        column_name = f'SN{i + 1}'
        level = generate_level(i)
        update_and_fill_sn(df_1, column_name, level)

    material_description_map = df_1.set_index('Material')['Material Description'].to_dict()
    sn_columns = [f'SN{i}' for i in range(1, n)]
    dsn_columns = [f'DSN{i}' for i in range(1, n)]

    def fill_dsn_columns(df_1, sn_columns, dsn_columns, material_description_map):
        for sn_col, dsn_col in zip(sn_columns, dsn_columns):
            df_1.loc[:, dsn_col] = df_1[sn_col].map(material_description_map)

    fill_dsn_columns(df_1, sn_columns, dsn_columns, material_description_map)

    def clean_table(row):
        if row['Level'] == '*1':
            row['SN2'] = '(blank)'
            row['DSN2'] = '(blank)'
            row['SN3'] = '(blank)'
            row['DSN3'] = '(blank)'
        elif row['Level'] == '**2':
            row['SN3'] = '(blank)'
            row['DSN3'] = '(blank)'
        return row

    df_1 = df_1.apply(clean_table, axis=1)
    filtered_df = df_1.dropna(subset=['SN2', 'DSN2', 'SN3', 'DSN3'], how='all')
    columns_to_keep = ['SN1', 'DSN1', 'SN2', 'DSN2', 'SN3', 'DSN3', 'PN']
    filtered_df = filtered_df[columns_to_keep]

    grouped = filtered_df.groupby(['SN1', 'DSN1', 'SN2', 'DSN2', 'SN3', 'DSN3', 'PN']).size().reset_index(name='count')
    pivot_table = grouped.pivot(index=['SN1', 'DSN1', 'SN2', 'DSN2', 'SN3', 'DSN3'], columns='PN', values='count').fillna(0).reset_index()
    pivot_table.replace(0, np.nan, inplace=True)
    pivot_table['Total'] = pivot_table.sum(axis=1, numeric_only=True)

    # Ajouter la colonne DS Général aux différentes feuilles
    def add_ds_general_column(sheet_name, df):
        if sheet_name == 'FIL SIMPLE':
            df['DS Général'] = df['SN1'].apply(lambda x: generate_ds_general(x, df_mmsta))
        elif sheet_name in ['joint', 'double', 'twist', 'SQUIB', 'GW']:
            df['DS Général'] = df['SN2'].apply(lambda x: generate_ds_general(x, df_mmsta))
        elif sheet_name == 'super group':
            df['DS Général'] = df['SN3'].apply(lambda x: generate_ds_general(x, df_mmsta))
        return df

    filters = {
        'FIL SIMPLE': 'circuit ',
        'double': 'double',
        'twist': 'twisted',
        'joint': 'joint',
        'super group': 'super group',
        'SQUIB': 'simple super group',
        'cut tube': 'cut tube|GAFT',
        'GW': 'group wire'
    }

    with pd.ExcelWriter("MMSTA_separe.xlsx", engine='openpyxl') as writer:
        df_mmsta.to_excel(writer, sheet_name='MMSTA', index=False)
        pivot_table.to_excel(writer, sheet_name='SEPARER', index=False)

        for sheet_name, filter_str in filters.items():
            if sheet_name == 'super group':
                filtered_df = pivot_table[pivot_table['DSN1'].str.contains(filter_str, case=False, na=False) &
                                          ~pivot_table['DSN1'].str.contains('simple super group', case=False, na=False)]
            else:
                filtered_df = pivot_table[pivot_table['DSN1'].str.contains(filter_str, case=False, na=False)]

            filtered_df = add_ds_general_column(sheet_name, filtered_df)
            filtered_df.to_excel(writer, sheet_name=sheet_name, index=False)

    workbook = openpyxl.load_workbook("MMSTA_separe.xlsx")
    sheet = workbook['SEPARER']
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    for cell in sheet[1]:
        cell.fill = yellow_fill
    workbook.save("MMSTA_separe.xlsx")

    print(f"Le fichier Excel généré: MMSTA_separe.xlsx")
    return "MMSTA_separe.xlsx"

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python separation.py <mmsta_file_path>")
        sys.exit(1)
    
    mmsta_file = sys.argv[1]
    output_file = separate_files(mmsta_file)
    print(f"Fichier Excel généré: {output_file}")
